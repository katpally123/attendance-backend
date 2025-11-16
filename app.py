import os
from io import BytesIO
from typing import Dict, Any

from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

TEMPLATE_FILE = "Site_Split_Template.xlsx"

# Final row mapping
ROW_MAP = {
    "RegularHC": 6,
    "PresentHC": 7,
    "SwapOut": 8,
    "SwapExpected": 9,
    "SwapPresent": 10,
    "VTO": 12,
    "VETAccepted": 13,
    "VETPresent": 14,
    "METExpected": 18,
    "METPresent": 19,
}

# Final column mapping
COL_MAP = {
    "inbound_amzn": "B",
    "inbound_temp": "C",
    "da_amzn": "D",
    "da_temp": "E",
    "icqa_amzn": "G",
    "icqa_temp": "H",
    "crets_amzn": "I",
    "crets_temp": "J",
}


def resolve_template_path(filename: str) -> str:
    """Return absolute path to template, relative to this file's directory."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, filename)


def validate_payload(payload: Any) -> Dict[str, Dict[str, int]]:
    """
    Validate and normalize the incoming payload.
    - Must be a dict of metric -> dict of department -> int
    - Unknown metrics are ignored; unknown departments are ignored
    - Non-int values are coerced to int when possible, else default to 0
    """
    if not isinstance(payload, dict):
        raise ValueError("Payload must be a JSON object at the top level")

    allowed_metrics = set(ROW_MAP.keys())
    allowed_depts = set([
        "inbound_amzn",
        "inbound_temp",
        "da_amzn",
        "da_temp",
        "icqa_amzn",
        "icqa_temp",
        "crets_amzn",
        "crets_temp",
    ])

    normalized: Dict[str, Dict[str, int]] = {}
    for metric, dept_data in payload.items():
        if metric not in allowed_metrics:
            continue
        safe_dept_data: Dict[str, int] = {}
        if isinstance(dept_data, dict):
            for dept, val in dept_data.items():
                if dept not in allowed_depts:
                    continue
                try:
                    safe_dept_data[dept] = int(val)
                except Exception:
                    safe_dept_data[dept] = 0
        normalized[metric] = safe_dept_data
    return normalized


def fill_sheet(ws, data: Dict[str, Dict[str, int]]):
    """
    Fill the DD-Metrics sheet with the given data dict.
    """
    for metric, row in ROW_MAP.items():
        metric_data = data.get(metric, {}) or {}

        # MET rows are forced to zero
        if metric.startswith("MET"):
            inbound_amzn = inbound_temp = da_amzn = da_temp = 0
            icqa_amzn = icqa_temp = 0
            crets_amzn = crets_temp = 0

            # write zeros into all dept cells
            for col in ["B", "C", "D", "E", "G", "H", "I", "J"]:
                ws[f"{col}{row}"] = 0

        else:
            # normal metrics use provided values
            inbound_amzn = metric_data.get("inbound_amzn", 0)
            inbound_temp = metric_data.get("inbound_temp", 0)
            da_amzn = metric_data.get("da_amzn", 0)
            da_temp = metric_data.get("da_temp", 0)
            icqa_amzn = metric_data.get("icqa_amzn", 0)
            icqa_temp = metric_data.get("icqa_temp", 0)
            crets_amzn = metric_data.get("crets_amzn", 0)
            crets_temp = metric_data.get("crets_temp", 0)

            # write individual dept values
            ws[f"B{row}"] = inbound_amzn
            ws[f"C{row}"] = inbound_temp
            ws[f"D{row}"] = da_amzn
            ws[f"E{row}"] = da_temp
            ws[f"G{row}"] = icqa_amzn
            ws[f"H{row}"] = icqa_temp
            ws[f"I{row}"] = crets_amzn
            ws[f"J{row}"] = crets_temp

        # SDC TOTAL = inbound + DA  → col F
        sdc_total = inbound_amzn + inbound_temp + da_amzn + da_temp
        ws[f"F{row}"] = sdc_total

        # IXD TOTAL = CRETs only → col K
        ixd_total = crets_amzn + crets_temp
        ws[f"K{row}"] = ixd_total

        # Grand TOTAL = SDC + ICQA + CRETs → col L
        grand_total = sdc_total + icqa_amzn + icqa_temp + ixd_total
        ws[f"L{row}"] = grand_total


@app.route("/api/generate-dashboard", methods=["POST"])
def generate_dashboard():
    """Generate the Excel using a JSON payload, return as a file download."""
    if not request.is_json:
        return jsonify({"error": "Content-Type must be application/json"}), 415

    raw_payload = request.get_json(silent=True)
    if raw_payload is None:
        return jsonify({"error": "Invalid or missing JSON payload"}), 400

    try:
        payload = validate_payload(raw_payload)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    try:
        template_path = resolve_template_path(TEMPLATE_FILE)
        wb = load_workbook(template_path)
    except FileNotFoundError:
        return (
            jsonify({
                "error": "Template file not found",
                "details": f"Missing '{TEMPLATE_FILE}' next to app.py",
            }),
            400,
        )
    except Exception as exc:
        return jsonify({"error": "Failed to load template", "details": str(exc)}), 500

    ws = wb["DD-Metrics"] if "DD-Metrics" in wb.sheetnames else wb.active

    try:
        fill_sheet(ws, payload)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
    except Exception as exc:
        return jsonify({"error": "Failed to generate workbook", "details": str(exc)}), 500

    return send_file(
        output,
        as_attachment=True,
        download_name="Daily_Attendance_Auto.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/generate-dashboard/test", methods=["GET"])
def generate_dashboard_test():
    """Convenience endpoint to generate a file from built-in dummy data (manual testing)."""
    dummy = {
        "RegularHC": {
            "inbound_amzn": 10,
            "inbound_temp": 5,
            "da_amzn": 8,
            "da_temp": 3,
            "icqa_amzn": 4,
            "icqa_temp": 1,
            "crets_amzn": 9,
            "crets_temp": 6,
        },
        "PresentHC": {
            "inbound_amzn": 9,
            "inbound_temp": 5,
            "da_amzn": 6,
            "da_temp": 2,
            "icqa_amzn": 4,
            "icqa_temp": 1,
            "crets_amzn": 8,
            "crets_temp": 6,
        },
        "SwapOut": {},
        "SwapExpected": {},
        "SwapPresent": {},
        "VTO": {},
        "VETAccepted": {},
        "VETPresent": {},
        "METExpected": {},
        "METPresent": {},
    }

    try:
        template_path = resolve_template_path(TEMPLATE_FILE)
        wb = load_workbook(template_path)
    except Exception as exc:
        return jsonify({"error": "Failed to load template", "details": str(exc)}), 500

    ws = wb["DD-Metrics"] if "DD-Metrics" in wb.sheetnames else wb.active
    try:
        fill_sheet(ws, dummy)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
    except Exception as exc:
        return jsonify({"error": "Failed to generate workbook", "details": str(exc)}), 500

    return send_file(
        output,
        as_attachment=True,
        download_name="Daily_Attendance_Auto.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/")
def home():
    return "Backend running!"


if __name__ == "__main__":
    # Render supplies PORT; default to 10000 for local dev
    port = int(os.environ.get("PORT", "10000"))
    debug = os.environ.get("FLASK_DEBUG", "").lower() in ("1", "true", "yes")
    app.run(host="0.0.0.0", port=port, debug=debug)
